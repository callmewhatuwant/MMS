#SQL Import
from multiprocessing.dummy import active_children
from operator import truediv
import mysql.connector
import re
import openpyxl 
wb = openpyxl.load_workbook(filename = 'IstAufwand_MitteAug_TD_2.xlsx')
sheet_obj = wb.active
m_row = sheet_obj.max_row


mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="123456",
  database="sOliver"
)

for i in range(1, m_row + 1):
    
    TicketID = ""
    Kommentar = ""
    Vorgang = ""
    
    cell_obj3 = sheet_obj.cell(row = i, column = 3)
    print(cell_obj3.value)
    cell_obj4 = sheet_obj.cell(row = i, column = 4)
    print(cell_obj4.value)
    cell_obj6 = sheet_obj.cell(row = i, column = 6)
    print(cell_obj6.value)
    
    cell_object9 = sheet_obj.cell(row = i, column = 9)
    match cell_object9.value:
      case "01.01_Fachliche und Technische Analyse":
        Vorgang = "Analyse"
      case "02.01_Entwicklung":
        Vorgang = "Entwicklung"
      case "02.04_Bugfixing":
        Vorgang = "Bugfix"
      case "03.01_Organisatorisches Projektmanagement":
        Vorgang = "PO/PL"
      case "03.02_Technisches Projektmanagement":
        Vorgang = "TP"
      case "03.04_Testautomatisierung (regelmäßige Aufwände)":
        Vorgang = "TA"
      case "03.05_Regressionstestset WebShop Platform":
        Vorgang = "RG"
      case "03.07_Solution_Support":
        Vorgang = "solSup"
      case "03.10_Build&Deployment releaseabhängig":
        Vorgang = "BnD_release"
      case "03.12_Scrum Master":
        Vorgang = "Scrum"
      case "03.13_Sprintmeetings":
        Vorgang = "Meeting"
      case "02.03_Testing & QA":
        Vorgang = "Test"
      case "02.02_Dokumentation":
        Vorgang = "Doku"
      case "03.05_Regressionstestset WebShop Platform (für alle Sites)":
        Vorgang = "TA"
      case "Last-Level-Support_nicht_verrechenbar":
        Vorgang = "LLS nicht v"
      case "03.08_sonstige nicht verrechenbare Zeiten (mit Kommentar)":
        Vorgang = "andere Z"
      case "03.11_Build&Deployment releaseunabhängig":
        Vorgang = "BnD_nonrelease"
        print(Vorgang)
      
        
    cell_obj11 = sheet_obj.cell(row = i, column = 11)
    x = re.search ("^SOLI-[0-9][0-9][0-9][0-9][0-9]$",cell_obj11.value)
    if x:
      TicketID = cell_obj11.value
      print (TicketID)
    else:
      Kommentar = cell_obj11.value   
      print(Kommentar)
          
    mycursor = mydb.cursor()
    sql = "INSERT INTO sOliver (Datum, Aufwand, Mitarbeiter, Vorgang, Kommentar, TicketID) VALUES (%s, %s, %s, %s, %s, %s)"
    val = (cell_obj3.value, cell_obj4.value, cell_obj6.value, Vorgang, Kommentar, TicketID)
    mycursor.execute(sql, val)

mydb.commit()

print(mycursor.rowcount, "record inserted.")
