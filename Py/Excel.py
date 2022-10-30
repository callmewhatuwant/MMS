import re
import openpyxl
ref_workbook= openpyxl.load_workbook('excel.xlsx')
wb = ref_workbook
# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells

#Input
#ws['A1'] = ("Datum")
#ws['B1'] = ("Anfangszeit")
#ws['C1'] = ("Aufgabe")
#ws['D1'] = ("Endzeit")

Datum = input("Datum angeben: ")
Anfangszeit = input("Anfangszeit angeben: ")
Aufgabe = input("Aufgabe beschreiben: ")
Endzeit = input("Endzeit angeben: ")

ws.append([Datum, Anfangszeit, Aufgabe, Endzeit])

#Datum
if len (Datum) != 10:
  print("Bitte korrigieren: dd.mm.jjjj ")
  quit()

x = re.search ("^[0-3][0-9]\.[0-1][0-2]\.[1-9][0-9][0-9][0-9]$", Datum)
if not x:
  print ("Bitte korrigieren: dd.mm.jjjj ")
  quit()

#Anfangszeit
if len (Anfangszeit) != 5 :
  print("Bitte korrigiere Anfangszeit:00:00 ")
  quit()

x = re.search ("^[0-2][0-9]\:[0-5][0-9]$", Anfangszeit)
if not x:
  print ("Bitte korrigiere Anfangszeit:00:00 ")
  quit()

#Aufgabe
if len (Aufgabe) <10 :
  print ("Bitte korrigiere: mehr als 10 Zeichen ")
  quit()

#Endzeit
if len (Endzeit) != 5:
  print ("Bitte korrigiere Endzeit:00:00 ")
  quit()

x = re.search ("^[0-2][0-9]\:[010.12p0-5][0-9]$", Endzeit)
if not x:
  print ("Bitte korrigiere Endzeit :00:00 ")
  quit()
  
  
  # Save the file
wb.save("excel.xlsx")
  

